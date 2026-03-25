#!/usr/bin/env python3
"""
Helium Hustle — Reverse data converter: JSON -> xlsx

Converts godot/data/*.json back into a spreadsheet that data/convert.py
can read. Use this when you want to visually inspect or edit game data:

    1. python data/json_to_xlsx.py          # JSON -> xlsx
    2. (edit the xlsx in Excel / Google Sheets)
    3. python data/convert.py               # xlsx -> JSON  (ground truth)

The JSON files are ground truth. The xlsx is a human-readable intermediate.
Do not treat the xlsx as the source of record.

Output: data/Helium Hustle Datasheets.xlsx  (overwrites if present)
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

DATA_DIR    = Path(__file__).parent
GODOT_DATA  = DATA_DIR.parent / "godot" / "data"


# ============================================================================
# Cell encoding helpers (Resources / Buildings / Commands)
# ============================================================================

def encode_requires(req: dict) -> str:
    if not req or req.get("type") == "none":
        return "none"
    return f"{req['type']}={req['value']}"


def encode_costs(costs: dict) -> list:
    return [f"{res}={val:g}" for res, val in costs.items()]


def encode_production(production: dict) -> list:
    return [f"{res}+{val:g}" for res, val in production.items()]


def encode_upkeep(upkeep: dict) -> list:
    return [f"{res}-{val:g}" for res, val in upkeep.items()]


def encode_building_effects(effects: list) -> list:
    """Encode building effects using prefix_resource+value format."""
    result = []
    for e in effects:
        if set(e.keys()) == {"effect"}:
            result.append(e["effect"])
        elif "prefix" in e:
            result.append(f"{e['prefix']}_{e['resource']}{e['operator']}{e['value']:g}")
        else:
            result.append(f"{e['resource']}{e['operator']}{e['value']:g}")
    return result


def encode_command_effect(e: dict) -> str:
    """
    Encode a command effect as 'effect_name key=val ...' string.

    Examples:
      {"effect": "boredom_add", "value": 0.04}  -> 'boredom_add value=0.04'
      {"effect": "launch_full_pads"}             -> 'launch_full_pads'
      {"effect": "overclock", "target": "extraction", "bonus": 0.05, "duration": 5}
                                                 -> 'overclock target=extraction bonus=0.05 duration=5'
    """
    parts = [e["effect"]]
    for k, v in e.items():
        if k == "effect":
            continue
        parts.append(f"{k}={v:g}" if isinstance(v, float) else f"{k}={v}")
    return " ".join(parts)


def pad(lst: list, n: int) -> list:
    return (lst + ["x"] * n)[:n]


# ============================================================================
# Config tab helpers
# ============================================================================

def _flatten_config(obj, prefix=''):
    """Yield (dotpath, value) pairs for all leaf nodes of a nested structure."""
    if isinstance(obj, dict):
        for k, v in obj.items():
            new_prefix = f"{prefix}.{k}" if prefix else k
            yield from _flatten_config(v, new_prefix)
    elif isinstance(obj, list) and obj and isinstance(obj[0], dict):
        for i, item in enumerate(obj):
            yield from _flatten_config(item, f"{prefix}[{i}]")
    elif isinstance(obj, list):
        yield prefix, ','.join(str(x) for x in obj)
    else:
        yield prefix, obj


# ============================================================================
# Sheet writers
# ============================================================================

def write_resources(ws, resources: list) -> None:
    ws.append(["Resource", "Short name", "Storage base"])
    for r in resources:
        ws.append([
            r["name"],
            r["short_name"],
            r["storage_base"] if r["storage_base"] is not None else "x",
        ])


def write_buildings(ws, buildings: list) -> None:
    ws.append([
        "Building", "Short name", "Category", "Requires", "Land", "Scaling",
        "Cost 1", "Cost 2", "Cost 3", "Cost 4",
        "Prod 1", "Prod 2", "Prod 3",
        "Upkeep 1", "Upkeep 2",
        "Effect 1", "Effect 2", "Effect 3",
        "Desc",
    ])
    for b in buildings:
        row = [
            b["name"],
            b["short_name"],
            b.get("category", ""),
            encode_requires(b.get("requires", {})),
            b["land"],
            b["cost_scaling"],
        ]
        row += pad(encode_costs(b.get("costs", {})), 4)
        row += pad(encode_production(b.get("production", {})), 3)
        row += pad(encode_upkeep(b.get("upkeep", {})), 2)
        row += pad(encode_building_effects(b.get("effects", [])), 3)
        row += [b.get("description") or "x"]
        ws.append(row)


def write_commands(ws, commands: list) -> None:
    ws.append([
        "Command", "Short name", "Requires",
        "Cost 1", "Cost 2", "Cost 3",
        "Prod 1", "Prod 2", "Prod 3",
        "Effect 1", "Effect 2", "Effect 3",
        "Desc",
    ])
    for c in commands:
        row = [
            c["name"],
            c["short_name"],
            encode_requires(c.get("requires", {})),
        ]
        row += pad(encode_costs(c.get("costs", {})), 3)
        row += pad(encode_production(c.get("production", {})), 3)
        row += pad([encode_command_effect(e) for e in c.get("effects", [])], 3)
        row += [c.get("description") or "x"]
        ws.append(row)


def write_config(ws, config: dict) -> None:
    ws.append(["Key", "Value"])
    for i, (section_key, section_val) in enumerate(config.items()):
        if i > 0:
            ws.append(["", ""])  # blank separator between top-level sections
        for key_path, value in _flatten_config({section_key: section_val}):
            ws.append([key_path, value])


def style_header(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9D9D9")


# ============================================================================
# Main
# ============================================================================

def main():
    required = ["resources.json", "buildings.json", "commands.json", "game_config.json"]
    missing = [name for name in required if not (GODOT_DATA / name).exists()]
    if missing:
        print(f"ERROR: Missing JSON files in {GODOT_DATA}:")
        for m in missing:
            print(f"  {m}")
        sys.exit(1)

    resources   = json.loads((GODOT_DATA / "resources.json").read_text(encoding="utf-8"))
    buildings   = json.loads((GODOT_DATA / "buildings.json").read_text(encoding="utf-8"))
    commands    = json.loads((GODOT_DATA / "commands.json").read_text(encoding="utf-8"))
    game_config = json.loads((GODOT_DATA / "game_config.json").read_text(encoding="utf-8"))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_r = wb.create_sheet("Resources")
    write_resources(ws_r, resources)
    style_header(ws_r)

    ws_b = wb.create_sheet("Buildings")
    write_buildings(ws_b, buildings)
    style_header(ws_b)

    ws_c = wb.create_sheet("Commands")
    write_commands(ws_c, commands)
    style_header(ws_c)

    ws_cfg = wb.create_sheet("Config")
    write_config(ws_cfg, game_config)
    style_header(ws_cfg)

    out = DATA_DIR / "Helium Hustle Datasheets.xlsx"
    wb.save(out)
    print(f"Written: {out}")
    print(f"  Resources: {len(resources)}")
    print(f"  Buildings: {len(buildings)}")
    print(f"  Commands:  {len(commands)}")
    print(f"  Config:    {sum(1 for _ in _flatten_config(game_config))} keys")
    print()
    print("Edit the xlsx, then run: python data/convert.py")


if __name__ == "__main__":
    main()
