#!/usr/bin/env python3
"""
Helium Hustle Datasheet Converter
Converts Helium Hustle Datasheets.xlsx into JSON files for Godot.

NOTE: The JSON files in godot/data/ are ground truth. This script converts
an xlsx (human-readable intermediate) back into those JSON files. To edit
data visually, use the round-trip workflow:
    python data/json_to_xlsx.py   # JSON -> xlsx for editing
    (edit xlsx)
    python data/convert.py        # xlsx -> JSON (ground truth)

Usage:
    python convert.py [path_to_xlsx]

If no path given, looks for "Helium Hustle Datasheets.xlsx" in the same directory.
Outputs to godot/data/ (four files: resources.json, buildings.json,
commands.json, game_config.json).

========================================================================
SPREADSHEET FORMAT SPEC
========================================================================

Four tabs: Resources, Buildings, Commands, Config.

CELL ENCODING (Resources / Buildings / Commands tabs):
  "x"                   = empty/null (ignored by parser)
  "shortname=amount"    = cost (one-time purchase price, operator =)
  "shortname+amount"    = production (per-tick gain, operator +)
  "shortname-amount"    = upkeep (per-tick drain, operator -)
  "prefix_shortname+amount" = effect (e.g. store_eng+50, load_he3+2)
  "flagname"            = flag effect with no value (e.g. next_cmd_double)

Operators are CONFIRMATORY: the parser validates that = appears in Cost
columns, + in Prod columns, - in Upkeep columns. Mismatches are errors.
Effect columns accept any operator or bare flags.

Resource short names are defined in the Resources tab.

SHEET SCHEMAS:
  Resources: Resource | Short name | Storage base
  Buildings: Building | Short name | Category | Requires | Land | Scaling |
             Cost 1-4 | Prod 1-3 | Upkeep 1-2 | Effect 1-3 | Desc
  Commands:  Command | Short name | Requires |
             Cost 1-3 | Prod 1-3 | Effect 1-3 | Desc
  Config:    Key | Value
             Flat dot-notation key paths, e.g.:
               starting_resources.eng        100
               boredom_curve[0].day          0
               shipment.fuel_per_pad         20
               ideology.rank_thresholds      70,175,333,570,925
             Rows with empty/null Key or Key starting with # are skipped.
             Comma-separated values are parsed as arrays of numbers.

Column caps (Cost 1-4, Prod 1-3, etc.) are intentional design limits.
Headers are normalized to lowercase+stripped, so casing doesn't matter.

REQUIRES COLUMN:
  "none"                = no prerequisite
  "building=short_name" = requires at least one of that building
  "research=short_name" = (future) requires a research unlock

EXTENDING THIS SCRIPT:
  - New resource: add row to Resources tab. No script changes needed.
  - New building/command: add row to respective tab. No script changes.
  - New config key: add row to Config tab. No script changes needed.
  - New column cap (e.g. Cost 5): update the range() in the converter
    function AND update the xlsx generation if applicable.
  - New effect prefix (e.g. "mult_"): no script changes needed, the
    prefix parser handles any prefix_resource+value pattern.
  - New requires type: add handling in parse_requires() if needed.
  - New sheet (e.g. Research): add a convert_research() function
    following the same pattern, add to converters list in main().
========================================================================
"""

import json
import re
import sys
from pathlib import Path
from typing import Optional

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)


# ============================================================================
# Cell parsing (Resources / Buildings / Commands tabs)
# ============================================================================

COLUMN_OPERATORS = {
    "cost": "=",
    "prod": "+",
    "upkeep": "-",
}


def parse_resource_cell(raw: str, col_type: Optional[str], context: str) -> Optional[dict]:
    """
    Parse a cell like 'cred=80', 'eng+4', 'eng-2', 'store_eng+50', 'load_he3+2',
    'next_cmd_double', or 'x'.

    Returns None for 'x'.
    Returns dict with keys: resource, operator, value, prefix (if any).
    """
    if raw is None or str(raw).strip().lower() == "x":
        return None

    raw = str(raw).strip()

    if col_type == "effect" and not any(op in raw for op in "=+-"):
        return {"effect": raw}

    match = re.match(r'^([a-z_]*?)([a-z][a-z0-9]*)([=+\-])(.+)$', raw)
    if not match:
        raise ValueError(f"{context}: Cannot parse cell '{raw}'")

    for op in "=+-":
        if op in raw:
            left, right = raw.split(op, 1)
            try:
                value = float(right)
            except ValueError:
                raise ValueError(f"{context}: Bad numeric value in '{raw}'")

            if col_type in COLUMN_OPERATORS:
                expected_op = COLUMN_OPERATORS[col_type]
                if op != expected_op:
                    raise ValueError(
                        f"{context}: Operator mismatch in '{raw}'. "
                        f"Column type '{col_type}' expects '{expected_op}', got '{op}'"
                    )

            if "_" in left and col_type == "effect":
                parts = left.rsplit("_", 1)
                return {"prefix": parts[0], "resource": parts[1], "operator": op, "value": value}
            else:
                return {"resource": left, "operator": op, "value": value}

    raise ValueError(f"{context}: No operator found in '{raw}'")


def parse_command_effect_cell(raw) -> Optional[dict]:
    """
    Parse a command effect cell using 'effect_name key=val ...' format.

    Examples:
      'boredom_add value=0.04'          -> {"effect": "boredom_add", "value": 0.04}
      'launch_full_pads'                -> {"effect": "launch_full_pads"}
      'overclock target=extraction bonus=0.05 duration=5'
                                        -> {"effect": "overclock", "target": "extraction",
                                            "bonus": 0.05, "duration": 5}
    Returns None for 'x' or empty cells.
    """
    if raw is None or str(raw).strip().lower() == "x":
        return None
    raw = str(raw).strip()
    parts = raw.split()
    result: dict = {"effect": parts[0]}
    for kv in parts[1:]:
        k, _, v = kv.partition("=")
        try:
            fv = float(v)
            result[k] = int(fv) if fv == int(fv) else fv
        except ValueError:
            result[k] = v
    return result


def parse_requires(raw: str) -> dict:
    if raw is None or str(raw).strip().lower() == "none":
        return {"type": "none"}
    raw = str(raw).strip()
    if "=" in raw:
        req_type, req_value = raw.split("=", 1)
        return {"type": req_type, "value": req_value}
    raise ValueError(f"Cannot parse Requires value: '{raw}'")


# ============================================================================
# Sheet reader
# ============================================================================

def read_sheet(wb, sheet_name: str) -> tuple:
    ws = wb[sheet_name]
    raw_headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    headers = [h.strip().lower() if h else "" for h in raw_headers]

    rows = []
    for r in range(2, ws.max_row + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None for v in values):
            continue
        rows.append(dict(zip(headers, values)))
    return headers, rows


def get(row, key, sheet_name, row_index, headers, required=True):
    if key in row:
        return row[key]
    if required:
        raise KeyError(
            f"\n  Sheet '{sheet_name}', row {row_index}: missing column '{key}'"
            f"\n  Available headers: {headers}"
            f"\n  Row data: {row}"
        )
    return None


# ============================================================================
# Resources / Buildings / Commands converters
# ============================================================================

def convert_resources(wb) -> list:
    headers, rows = read_sheet(wb, "Resources")
    resources = []
    for i, row in enumerate(rows, start=2):
        g = lambda key, req=True: get(row, key, "Resources", i, headers, req)
        storage = g("storage base")
        entry = {
            "name": g("resource"),
            "short_name": g("short name"),
            "storage_base": float(storage) if storage != "x" and storage is not None else None,
        }
        resources.append(entry)
    return resources


def convert_buildings(wb) -> list:
    headers, rows = read_sheet(wb, "Buildings")
    buildings = []

    for i, row in enumerate(rows, start=2):
        g = lambda key, req=True: get(row, key, "Buildings", i, headers, req)
        name = g("building")
        ctx = f"Buildings/{name}"

        costs = []
        for j in range(1, 5):
            parsed = parse_resource_cell(row.get(f"cost {j}"), "cost", f"{ctx}/Cost {j}")
            if parsed:
                costs.append(parsed)

        productions = []
        for j in range(1, 4):
            parsed = parse_resource_cell(row.get(f"prod {j}"), "prod", f"{ctx}/Prod {j}")
            if parsed:
                productions.append(parsed)

        upkeeps = []
        for j in range(1, 3):
            parsed = parse_resource_cell(row.get(f"upkeep {j}"), "upkeep", f"{ctx}/Upkeep {j}")
            if parsed:
                upkeeps.append(parsed)

        effects = []
        for j in range(1, 4):
            parsed = parse_resource_cell(row.get(f"effect {j}"), "effect", f"{ctx}/Effect {j}")
            if parsed:
                effects.append(parsed)

        entry = {
            "name": name,
            "short_name": g("short name"),
            "category": g("category", req=False) or "",
            "requires": parse_requires(g("requires", req=False)),
            "land": int(g("land")),
            "cost_scaling": float(g("scaling")),
            "costs": {c["resource"]: c["value"] for c in costs},
            "production": {p["resource"]: p["value"] for p in productions},
            "upkeep": {u["resource"]: abs(u["value"]) for u in upkeeps},
            "effects": effects,
            "description": g("desc", req=False) or "",
        }
        buildings.append(entry)

    return buildings


def convert_commands(wb) -> list:
    headers, rows = read_sheet(wb, "Commands")
    commands = []

    for i, row in enumerate(rows, start=2):
        g = lambda key, req=True: get(row, key, "Commands", i, headers, req)
        name = g("command")
        ctx = f"Commands/{name}"

        costs = []
        for j in range(1, 4):
            parsed = parse_resource_cell(row.get(f"cost {j}"), "cost", f"{ctx}/Cost {j}")
            if parsed:
                costs.append(parsed)

        productions = []
        for j in range(1, 4):
            parsed = parse_resource_cell(row.get(f"prod {j}"), "prod", f"{ctx}/Prod {j}")
            if parsed:
                productions.append(parsed)

        effects = []
        for j in range(1, 4):
            parsed = parse_command_effect_cell(row.get(f"effect {j}"))
            if parsed:
                effects.append(parsed)

        entry = {
            "name": name,
            "short_name": g("short name"),
            "requires": parse_requires(g("requires", req=False)),
            "costs": {c["resource"]: c["value"] for c in costs},
            "production": {p["resource"]: p["value"] for p in productions},
            "effects": effects,
            "description": g("desc", req=False) or "",
        }
        commands.append(entry)

    return commands


# ============================================================================
# Config tab converter  (Key | Value  dot-notation rows -> nested dict)
# ============================================================================

_PATH_PART = re.compile(r'([^\.\[\]]+)|\[(\d+)\]')


def _auto_type(value):
    """Convert a cell value to the appropriate Python type."""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if value is None:
        return None
    s = str(value).strip()
    if s.lower() == 'true':
        return True
    if s.lower() == 'false':
        return False
    # Comma-separated list of numbers (e.g. "70,175,333,570,925")
    if ',' in s:
        try:
            parts = [float(x.strip()) for x in s.split(',')]
            return [int(x) if x == int(x) else x for x in parts]
        except (ValueError, OverflowError):
            return s
    # Integer or float
    try:
        f = float(s)
        return int(f) if f == int(f) else f
    except ValueError:
        return s


def _set_config_path(root: dict, path: str, value) -> None:
    """Set a value at a dot/bracket path in a nested dict (mutates root)."""
    tokens = []
    for key_part, idx_part in _PATH_PART.findall(path):
        if key_part:
            tokens.append(('key', key_part))
        else:
            tokens.append(('index', int(idx_part)))

    obj = root
    for i, (typ, key) in enumerate(tokens[:-1]):
        next_typ, _ = tokens[i + 1]
        if typ == 'key':
            if key not in obj:
                obj[key] = [] if next_typ == 'index' else {}
            obj = obj[key]
        else:  # index
            while len(obj) <= key:
                obj.append({})
            obj = obj[key]

    final_typ, final_key = tokens[-1]
    typed_val = _auto_type(value)
    if final_typ == 'key':
        obj[final_key] = typed_val
    else:
        while len(obj) <= final_key:
            obj.append(None)
        obj[final_key] = typed_val


def convert_config(wb) -> dict:
    """Parse the Config tab (Key | Value rows) into a nested dict."""
    ws = wb["Config"]
    result = {}
    for r in range(2, ws.max_row + 1):
        key = ws.cell(row=r, column=1).value
        value = ws.cell(row=r, column=2).value
        if not key or str(key).strip().startswith('#'):
            continue
        _set_config_path(result, str(key).strip(), value)
    return result


# ============================================================================
# Main
# ============================================================================

def main():
    script_dir = Path(__file__).parent
    if len(sys.argv) > 1:
        xlsx_path = Path(sys.argv[1])
    else:
        xlsx_path = script_dir / "Helium Hustle Datasheets.xlsx"

    if not xlsx_path.exists():
        print(f"ERROR: File not found: {xlsx_path}")
        sys.exit(1)

    print(f"Reading: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    output_dir = script_dir.parent / "godot" / "data"
    output_dir.mkdir(exist_ok=True)

    errors = []

    converters = [
        ("resources.json",   lambda: convert_resources(wb)),
        ("buildings.json",   lambda: convert_buildings(wb)),
        ("commands.json",    lambda: convert_commands(wb)),
        ("game_config.json", lambda: convert_config(wb)),
    ]

    for filename, converter in converters:
        try:
            data = converter()
            out_path = output_dir / filename
            with open(out_path, "w", newline="\n") as f:
                json.dump(data, f, indent=2)
                f.write("\n")
            size = f"{len(data)} entries" if isinstance(data, list) else "ok"
            print(f"  Wrote {filename}: {size}")
        except (ValueError, KeyError) as e:
            errors.append(str(e))
            print(f"  ERROR in {filename}: {e}")

    if errors:
        print(f"\n{'='*60}")
        print(f"FAILED: {len(errors)} validation error(s):")
        for err in errors:
            print(f"  - {err}")
        print(f"{'='*60}")
        sys.exit(1)
    else:
        print(f"\nDone. Output in: {output_dir}")


if __name__ == "__main__":
    main()
